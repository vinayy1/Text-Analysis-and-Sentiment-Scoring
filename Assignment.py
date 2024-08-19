# 1: Extracting the files

import requests
from bs4 import BeautifulSoup
import openpyxl

def extract_article_text(url, filename):
  """
  Extracts article title and text from a URL and saves it to a file.

  Args:
      url (str): The URL of the article to extract.
      filename (str): The filename to save the extracted text.
  """
  try:
    response = requests.get(url)
    response.raise_for_status()  # Raise an exception for bad status codes

    soup = BeautifulSoup(response.content, 'html.parser')

    # Find elements containing article title and text (logic might need adjustment based on website structure)
    title_element = soup.find('h1', class_=lambda class_: class_ and 'title' in class_)  # Adjust selector for title
    text_elements = soup.find_all('p')  # Adjust selector for paragraphs within article body

    # Combine text content from paragraphs
    article_text = ""
    for element in text_elements:
      article_text += element.text.strip() + "\n"

    # Save extracted text
    with open(filename, 'w', encoding='utf-8') as f:
      if title_element:
        f.write(title_element.text.strip() + "\n\n")  # Write title followed by newline
      f.write(article_text)

    print(f"Extracted text from {url} and saved to {filename}")
  except requests.exceptions.RequestException as e:
    print(f"Error: Could not extract text from {url} - {e}")

def read_urls_from_xlsx(filename):
  """
  Reads a list of URLs from an Excel spreadsheet.

  Args:
      filename (str): The filename of the Excel spreadsheet.

  Returns:
      list: A list of URLs extracted from the spreadsheet.
  """
  urls = []
  try:
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active  # Assuming URLs are in the first sheet

    # Extract URLs from specific columns 
    for row in sheet.iter_rows(min_row=2):  # Skip header row (assuming row 1)
      url_id = row[0].value  # Assuming URL_ID is in the first column (index 0)
      url = row[1].value  # Assuming URL is in the second column (index 1)
      if url:
        # Create filename based on URL_ID
        filename = f"article_{url_id}.txt"
        urls.append((url, filename))  # Store URL and corresponding filename in a tuple

  except FileNotFoundError:
    print(f"Error: Excel file {filename} not found")
  except openpyxl.exceptions.OpenpyxlError as e:
    print(f"Error: Could not read URLs from {filename} - {e}")

  return urls

if __name__ == "__main__":
  
  url_list = read_urls_from_xlsx('Input.xlsx')

  for url, filename in url_list:
    extract_article_text(url, filename)




# 2: Merging stop words
    
import nltk
nltk.download('stopwords')

nltk.download('punkt')



def merge_stop_words(stopwords_folder, output_file_path):
  """
  Merges the content of all stop word list files in a folder into a single file.

  Args:
      stopwords_folder (str): Path to the folder containing stop word list files.
      output_file_path (str): Path to the output file where merged stop words will be saved.
  """

  all_stop_words = set()
  for filename in os.listdir(stopwords_folder):
    if filename.endswith('.txt'):
      stopwords_file_path = os.path.join(stopwords_folder, filename)
      try:
          if filename == 'StopWords_Currencies.txt':  # Check for specific filename
              with open(stopwords_file_path, 'r', encoding='ISO-8859-1') as f:
                  stop_words = set(line.strip() for line in f)
          else:
              # Use UTF-8 encoding for other files (assuming they're different)
              with open(stopwords_file_path, 'r', encoding='utf-8') as f:
                  stop_words = set(line.strip() for line in f)
          all_stop_words.update(stop_words)
      except UnicodeDecodeError:
          print(f"Error: Could not decode file {stopwords_file_path} even with the specified encoding.")

  # Write all stop words to the output file
  with open(output_file_path, 'w') as f:
    for word in all_stop_words:
      f.write(f'{word}\n')  # Write each word with a newline character

if __name__ == '__main__':
  # Define paths
  stopwords_folder = '/Users/ashish/Desktop/Blackcoffer Intern/StopWords'  
  output_file_path = os.path.join(stopwords_folder, 'StopWords.txt')

  merge_stop_words(stopwords_folder, output_file_path)

  print(f'Merged stop words saved to: {output_file_path}')




# 3: Clean Files
  
import os


def clean_text_file(input_file_path, output_file_path, stopwords_folder):
  """
  Cleans a text file by removing stop words from the provided stop words list.

  Args:
      input_file_path (str): Path to the input text file.
      output_file_path (str): Path to the output cleaned text file.
      stopwords_folder (str): Path to the folder containing the stop words list file (StopWords.txt).
  """

  # Load stop words from the specified file
  stopwords_file_path = os.path.join(stopwords_folder, 'StopWords.txt')
  with open(stopwords_file_path, 'r') as f:
      stop_words = set(line.strip() for line in f)

  # Read text from input file
  try:
      with open(input_file_path, 'r', encoding='utf-8') as f:
          text = f.read()
  except UnicodeDecodeError:
      print(f"Error: Could not decode file {input_file_path}. Please check the encoding.")
      return

  # Tokenize text into words
  words = word_tokenize(text.lower())  # Convert text to lowercase

  # Remove stop words
  cleaned_words = [word for word in words if word not in stop_words]

  # Join cleaned words back into text with space separator
  cleaned_text = ' '.join(cleaned_words)

  # Write cleaned text to output file
  with open(output_file_path, 'w') as f:
      f.write(cleaned_text)


if __name__ == '__main__':
  # Define paths
  text_files_folder = '/Users/ashish/Desktop/Blackcoffer Intern/MyTextFiles'
  stopwords_folder = '/Users/ashish/Desktop/Blackcoffer Intern/StopWords'
  cleaned_files_folder = '/Users/ashish/Desktop/Blackcoffer Intern/Cleaned'

  # Create the cleaned files folder if it doesn't exist
  os.makedirs(cleaned_files_folder, exist_ok=True)  # Create folder if needed

  # Loop through all files in the text files folder (assuming numbering starts from 2011)
  for i in range(2011, 2158):
      filename = f'article_bctech{i}.txt'
      input_file_path = os.path.join(text_files_folder, filename)
      output_file_path = os.path.join(cleaned_files_folder, filename)
      clean_text_file(input_file_path, output_file_path, stopwords_folder)

  print(f'Cleaned files saved to: {cleaned_files_folder}')




# 4: Positive and Negative Files
  
def load_stopwords(filename):
  """Loads stop words from a text file.

  Args:
      filename (str): The path to the stop words file.

  Returns:
      set: A set of stop words.
  """
  stopwords = set()
  with open(filename, 'r', encoding='utf-8') as f:
    for line in f:
      stopwords.add(line.strip().lower())  # Remove whitespace and convert to lowercase
  return stopwords

def load_positive_negative_words(filename, stopwords):
  """Loads positive and negative words from a text file, excluding stop words.

  Args:
      filename (str): The path to the positive/negative words file.
      stopwords (set): A set of stop words.

  Returns:
      list: A list of words from the positive/negative words file (excluding stop words).
  """
  words = []
  encoding = 'utf-8'  # Default encoding (can be adjusted based on the file)
  if filename.endswith('.txt'):  # Check if the filename ends with .txt
    if 'positive' in filename:  # Check if filename suggests positive words
      encoding = 'ascii'  # Use ASCII encoding for positive-words.txt
    else:
      encoding = 'iso-8859-1'  # Use ISO-8859-1 for negative-words.txt
  try:
    with open(filename, 'r', encoding=encoding) as f:
      for line in f:
        word = line.strip().lower()
        if word not in stopwords:  # Include only words not in stopwords
          words.append(word)
  except UnicodeDecodeError:
    print(f"Error decoding file: {filename}. Skipping...")
  return words

def save_list(word_list, filename):
  """Saves the list of words to a text file.

  Args:
      word_list (list): The list of words to save.
      filename (str): The path to the output file.
  """
  with open(filename, 'w') as f:
    for word in word_list:
      f.write(f"{word}\n")

# Define file paths 
stopwords_file = "/Users/ashish/Desktop/Blackcoffer Intern/StopWords/StopWords.txt"
positive_words_file = "/Users/ashish/Desktop/Blackcoffer Intern/MasterDictionary/positive-words.txt"
negative_words_file = "/Users/ashish/Desktop/Blackcoffer Intern/MasterDictionary/negative-words.txt"
output_file = "/Users/ashish/Desktop/Blackcoffer Intern/MasterDictionary/all_words.txt"  # Updated output filename

# Load stop words
stopwords = load_stopwords(stopwords_file)

# Load positive and negative words, excluding stop words
positive_words = load_positive_negative_words(positive_words_file, stopwords)
negative_words = load_positive_negative_words(negative_words_file, stopwords)

# Combine positive and negative words (excluding duplicates)
all_words = list(set(positive_words + negative_words))  # Combine and remove duplicates

# Save the list of words to a text file
save_list(all_words, output_file)

print(f"All words from positive and negative files (excluding stop words) saved to: {output_file}")


# 5: Positive Dictionary and Negative Dictionary

def load_stopwords(filename):
  """Loads stop words from a text file.

  Args:
      filename (str): The path to the stop words file.

  Returns:
      set: A set of stop words.
  """
  stopwords = set()
  with open(filename, 'r', encoding='utf-8') as f:
    for line in f:
      stopwords.add(line.strip().lower())  # Remove whitespace and convert to lowercase
  return stopwords

def load_positive_negative_words(filename, stopwords):
  """Loads positive and negative words from a text file, excluding stop words.

  Args:
      filename (str): The path to the positive/negative words file.
      stopwords (set): A set of stop words.

  Returns:
      dict: A dictionary containing positive/negative words as keys and empty lists as values.
  """
  words = {'positive': []} if 'positive' in filename else {'negative': []}  # Create dictionary with appropriate key
  encoding = 'utf-8'  # Default encoding (can be adjusted based on the file)
  if filename.endswith('.txt'):  # Check if the filename ends with .txt
    if 'positive' in filename:  # Check if filename suggests positive words
      encoding = 'ascii'  # Use ASCII encoding for positive-words.txt
    else:
      encoding = 'iso-8859-1'  # Use ISO-8859-1 for negative-words.txt
  try:
    with open(filename, 'r', encoding=encoding) as f:
      for line in f:
        word = line.strip().lower()
        if word not in stopwords:  # Include only words not in stopwords
          words[list(words.keys())[0]].append(word)  # Append word to the existing key
  except UnicodeDecodeError:
    print(f"Error decoding file: {filename}. Skipping...")
  return words

def save_dictionary(dictionary, filename):
  """Saves the dictionary to a text file.

  Args:
      dictionary (dict): The dictionary to save.
      filename (str): The path to the output file.
  """
  with open(filename, 'w') as f:
    for sentiment, word_list in dictionary.items():
      f.write(f"{sentiment}:\n")
      for word in word_list:
        f.write(f"\t{word}\n")

# Define file paths (replace with your actual locations)
stopwords_file = "/Users/ashish/Desktop/Blackcoffer Intern/StopWords/StopWords.txt"
positive_words_file = "/Users/ashish/Desktop/Blackcoffer Intern/MasterDictionary/positive-words.txt"
negative_words_file = "/Users/ashish/Desktop/Blackcoffer Intern/MasterDictionary/negative-words.txt"
positive_dict_file = "/Users/ashish/Desktop/Blackcoffer Intern/MasterDictionary/positive_dictionary.txt"
negative_dict_file = "/Users/ashish/Desktop/Blackcoffer Intern/MasterDictionary/negative_dictionary.txt"

# Load stop words
stopwords = load_stopwords(stopwords_file)

# Load positive and negative words, excluding stop words
positive_words = load_positive_negative_words(positive_words_file, stopwords)
negative_words = load_positive_negative_words(negative_words_file, stopwords)

# Save positive and negative words to separate dictionary files
save_dictionary(positive_words, positive_dict_file)
save_dictionary(negative_words, negative_dict_file)

print(f"Positive words dictionary saved to: {positive_dict_file}")
print(f"Negative words dictionary saved to: {negative_dict_file}")


# 6: Calculating and writing scores to excel

import nltk
import openpyxl  # For interacting with Excel files

# Load positive and negative dictionaries
positive_dict = {}
negative_dict = {}
with open("/Users/ashish/Desktop/Blackcoffer Intern/MasterDictionary/positive_dictionary.txt", 'r') as f:
  for line in f:
    word = line.strip()
    positive_dict[word] = 1  # Assign value 1 for positive words
with open("/Users/ashish/Desktop/Blackcoffer Intern/MasterDictionary/negative_dictionary.txt", 'r') as f:
  for line in f:
    word = line.strip()
    negative_dict[word] = -1  # Assign value -1 for negative words

# Download required NLTK resources (comment out if already downloaded)
nltk.download('punkt')  # Download sentence tokenizer

def process_article(filename):
  """
  Processes an article by calculating positive score, negative score, polarity score, and subjectivity score.

  Args:
      filename (str): Path to the article text file.

  Returns:
      dict: A dictionary containing the calculated scores and filename.
  """
  with open(filename, 'r') as f:
    text = f.read()

  # Tokenize the text
  tokens = word_tokenize(text.lower())  # Lowercase and tokenize

  # Initialize scores
  positive_score = 0
  negative_score = 0

  # Calculate positive and negative scores
  for token in tokens:
    if token in positive_dict:
      positive_score += positive_dict[token]
    elif token in negative_dict:
      negative_score += negative_dict[token]

  # Calculate polarity score (avoid division by zero)
  total_score = positive_score + negative_score
  polarity_score = (positive_score - negative_score) / (total_score + 0.000001)

  # Clipping polarity score to -1 to 1
  # polarity_score = max(-1.0, min(1.0, polarity_score))

  # Calculate subjectivity score
  subjectivity_score = (positive_score + negative_score) / (len(tokens) + 0.000001)
    
    
    
  return {
    'filename': filename,
    'positive_score': positive_score,
    'negative_score': abs(negative_score),  # Make negative score positive
    'polarity_score': polarity_score,
    'subjectivity_score': subjectivity_score
  }

# Process all articles
article_scores = []
for i in range(2011, 2158):  # Loop through article numbers (adjust if needed)
  filename = f"/Users/ashish/Desktop/Blackcoffer Intern/Cleaned/article_bctech{i}.txt"
  article_scores.append(process_article(filename))

# Update Excel file
output_file = "/Users/ashish/Desktop/Blackcoffer Intern/Output Data Structure.xlsx"
wb = openpyxl.load_workbook(output_file)
sheet = wb.active

# Start writing data from the second row (assuming headers are in the first row)
row_index = 2
for article in article_scores:
  file_id = article['filename'].split("/")[-1].split(".")[0]  # Extract article ID from filename
  url = sheet.cell(row=row_index, column=2).value  # assuming URL is in column 2
  positive_score = article['positive_score']
  negative_score = article['negative_score']
  polarity_score = article['polarity_score']
  subjectivity_score = article['subjectivity_score']

  sheet.cell(row=row_index, column=1).value = file_id  # Update URL_ID
  sheet.cell(row=row_index, column=3).value = positive_score
  sheet.cell(row=row_index, column=4).value = negative_score
  sheet.cell(row=row_index, column=5).value = polarity_score
  sheet.cell(row=row_index, column=6).value = subjectivity_score

  row_index += 1

wb.save(output_file)
print(f"Scores written to Excel file: {output_file}")


# 7: Writing remaining scores to excel sheet

import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
import openpyxl
from nltk.corpus import stopwords  # Import stopwords class
import re  # Import for regular expressions


# Load stopwords
stop_words = set(stopwords.words('english'))


def count_syllables(word):
  """
  Estimates the number of syllables in a word using a simple rule.

  Args:
      word (str): The word to analyze.

  Returns:
      int: The estimated number of syllables.
  """
  vowels = 'aeiouAEIOU'
  count = 0
  i = 0
  while i < len(word):
    if word[i] in vowels and (i == 0 or word[i - 1] not in vowels):
      count += 1
    i += 1

  # Exceptions for common cases (e.g., "science", "written")
  if word.endswith("e") and word[:-1] not in vowels:
    count -= 1
  if word.endswith("ing") and word[:-3] not in vowels:
    count -= 1

  return count


def calculate_readability(filename):
  """
  Calculates readability scores for a text file.

  Args:
      filename (str): Path to the text file.

  Returns:
      dict: A dictionary containing the calculated scores.
  """
  with open(filename, 'r') as f:
    text = f.read()

  # Tokenize sentences and words (lowercase)
  sentences = sent_tokenize(text.lower())
  words = word_tokenize(text.lower())

  # Remove stop words and punctuations
  filtered_words = [word for word in words if word not in stop_words and word.isalpha()]

  # Average Sentence Length
  avg_sentence_length = len(words) / len(sentences) if len(sentences) > 0 else 0.0

  # Complex Word Count (assuming words with 3+ syllables are complex)
  complex_word_count = sum(1 for word in words if count_syllables(word) >= 3)

  # Percentage of Complex Words
  percentage_complex_words = (complex_word_count / len(words)) * 100 if len(words) > 0 else 0.0

  # Fog Index
  fog_index = 0.4 * (avg_sentence_length + percentage_complex_words)

  return {
      'avg_sentence_length': avg_sentence_length,
      'percentage_complex_words': percentage_complex_words,
      'fog_index': fog_index
  }


def process_articles(folder_path, output_file):
  """
  Processes all articles in a folder, calculates readability scores, and writes them to Excel.

  Args:
      folder_path (str): Path to the folder containing articles.
      output_file (str): Path to the Excel output file.
  """
  # Load Excel workbook (assuming the file exists)
  wb = openpyxl.load_workbook(output_file)
  sheet = wb.active

  # Start from row 2 (assuming headers are in row 1)
  row_index = 2
  for i in range(2011, 2158):
    filename = f"{folder_path}/article_bctech{i}.txt"
    scores = calculate_readability(filename)

    # Update Excel sheet
    file_id = sheet.cell(row=row_index, column=1).value  # Assuming URL_ID is in column 1
    sheet.cell(row=row_index, column=7).value = scores['avg_sentence_length']  # Update Average Sentence Length
    sheet.cell(row=row_index, column=8).value = scores['percentage_complex_words']  # Update Percentage of Complex Words
    sheet.cell(row=row_index, column=9).value = scores['fog_index']  # Update Fog Index

    row_index += 1

  # Save the updated workbook
  wb.save(output_file)
  print(f"Readability scores written to Excel file: {output_file}")


# Set folder paths and file paths
folder_path = "/Users/ashish/Desktop/Blackcoffer Intern/MyTextFiles"
